import sqlite3
import csv
import os
import shutil
from datetime import datetime
from werkzeug.security import generate_password_hash

DB_PATH = os.environ.get("DB_PATH", os.path.join(os.path.dirname(__file__), "..", "data", "dongbu.db"))
SEED_CSV = os.path.join(os.path.dirname(__file__), "seed_empleados.csv")


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def init_db():
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS empleados (
            codigo TEXT PRIMARY KEY,
            nombre TEXT NOT NULL,
            cargo TEXT,
            fecha_ingreso TEXT,
            salario_mensual REAL,
            activo INTEGER DEFAULT 1,
            dui TEXT,
            correo TEXT,
            telefono TEXT,
            fecha_fin TEXT,
            salario_letras TEXT,
            isss TEXT,
            afp TEXT,
            isr TEXT,
            prestamo_personal TEXT,
            prestamo_bancario TEXT,
            fsv TEXT
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS registros (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            codigo_empleado TEXT NOT NULL,
            tipo TEXT NOT NULL CHECK(tipo IN ('vacacion','incapacidad')),
            fecha_inicio TEXT NOT NULL,
            fecha_fin TEXT NOT NULL,
            dias INTEGER NOT NULL,
            foto_path TEXT,
            observaciones TEXT,
            fecha_registro TEXT NOT NULL,
            estado TEXT DEFAULT 'pendiente' CHECK(estado IN ('pendiente', 'aprobado', 'rechazado')),
            FOREIGN KEY(codigo_empleado) REFERENCES empleados(codigo)
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS admin_users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            nombre_completo TEXT
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS templates (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            doc_type TEXT NOT NULL,
            filename TEXT NOT NULL,
            stored_path TEXT NOT NULL,
            is_active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS generated_documents (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            codigo_empleado TEXT NOT NULL,
            doc_type TEXT NOT NULL,
            docx_path TEXT NOT NULL,
            pdf_path TEXT,
            created_at TEXT NOT NULL,
            FOREIGN KEY(codigo_empleado) REFERENCES empleados(codigo)
        )
    """)

    def safe_execute_alter(query):
        try:
            cur.execute(query)
        except sqlite3.OperationalError as e:
            err_msg = str(e).lower()
            if "duplicate column name" in err_msg or "already exists" in err_msg:
                pass
            else:
                raise e

    # Dynamic migration to add 'estado' to 'registros' if it doesn't exist yet
    cur.execute("PRAGMA table_info(registros)")
    columns = [row[1] for row in cur.fetchall()]
    if "estado" not in columns:
        safe_execute_alter("ALTER TABLE registros ADD COLUMN estado TEXT DEFAULT 'pendiente' CHECK(estado IN ('pendiente', 'aprobado', 'rechazado'))")

    # Dynamic migration to add columns to 'empleados' if they don't exist yet
    cur.execute("PRAGMA table_info(empleados)")
    emp_columns = [row[1] for row in cur.fetchall()]
    
    migrations = [
        ("dui", "TEXT"),
        ("correo", "TEXT"),
        ("telefono", "TEXT"),
        ("fecha_fin", "TEXT"),
        ("salario_letras", "TEXT"),
        ("isss", "TEXT"),
        ("afp", "TEXT"),
        ("isr", "TEXT"),
        ("prestamo_personal", "TEXT"),
        ("prestamo_bancario", "TEXT"),
        ("fsv", "TEXT"),
        ("departamento", "TEXT"),
        ("fecha_contratacion", "TEXT"),
        ("dias_vacacion_pendientes", "REAL DEFAULT 0")
    ]
    for col_name, col_type in migrations:
        if col_name not in emp_columns:
            safe_execute_alter(f"ALTER TABLE empleados ADD COLUMN {col_name} {col_type}")

    # Dynamic migration to add 'generado_por' to 'generated_documents' if it doesn't exist yet
    cur.execute("PRAGMA table_info(generated_documents)")
    doc_columns = [row[1] for row in cur.fetchall()]
    if "generado_por" not in doc_columns:
        safe_execute_alter("ALTER TABLE generated_documents ADD COLUMN generado_por TEXT DEFAULT 'admin'")

    conn.commit()

    # Seed/update employees from seed_mvp_employees.json
    # Supports two formats:
    #   - Old API format: employee_code, full_name, job_title, hire_date, salary, ...
    #   - New SQLite export format: codigo, nombre, cargo, fecha_ingreso, salario_mensual, dui, ...
    mvp_json_path = os.path.join(os.path.dirname(__file__), "seed_mvp_employees.json")
    if os.path.exists(mvp_json_path):
        try:
            import json
            with open(mvp_json_path, "r", encoding="utf-8") as f:
                employees = json.load(f)

            imported_count = 0
            for emp in employees:
                try:
                    # ── Código ──────────────────────────────────────────────
                    raw_code = (emp.get("codigo")
                                or emp.get("employee_code")
                                or emp.get("id"))
                    if not raw_code:
                        continue
                    raw_code_str = str(raw_code).strip()
                    if raw_code_str.isdigit():
                        codigo = raw_code_str.zfill(4)
                    elif isinstance(raw_code, float):
                        codigo = str(int(raw_code)).zfill(4)
                    else:
                        codigo = raw_code_str  # e.g. "EMP-044485307"

                    # ── Nombre ───────────────────────────────────────────────
                    nombre = str(emp.get("nombre") or emp.get("full_name") or "").strip()
                    if not nombre:
                        continue

                    # ── DUI ──────────────────────────────────────────────────
                    dui_raw = emp.get("dui") or ""
                    dui = str(dui_raw).strip() if dui_raw else None

                    # ── Cargo ────────────────────────────────────────────────
                    cargo_raw = emp.get("cargo") or emp.get("job_title") or ""
                    cargo = str(cargo_raw).strip() if cargo_raw else None

                    # ── Fecha ingreso ────────────────────────────────────────
                    ingreso_raw = emp.get("fecha_ingreso") or emp.get("hire_date")
                    ingreso = str(ingreso_raw).strip()[:10] if ingreso_raw else None

                    # ── Fecha fin ────────────────────────────────────────────
                    fin_raw = emp.get("fecha_fin") or emp.get("end_date")
                    fin = str(fin_raw).strip()[:10] if fin_raw else None

                    # ── Fecha contratación ───────────────────────────────────
                    contratacion_raw = emp.get("fecha_contratacion")
                    contratacion = str(contratacion_raw).strip()[:10] if contratacion_raw else ingreso

                    # ── Salario ──────────────────────────────────────────────
                    salario_raw = emp.get("salario_mensual") or emp.get("salary")
                    if salario_raw:
                        cleaned_sal = "".join(c for c in str(salario_raw) if c.isdigit() or c == '.')
                        salario = float(cleaned_sal) if cleaned_sal else 0.0
                    else:
                        salario = 0.0

                    # ── Deducciones ──────────────────────────────────────────
                    salario_letras = str(emp.get("salario_letras") or emp.get("salary_words") or "").strip() or None
                    isss   = str(emp.get("isss")   or "").strip() or None
                    afp    = str(emp.get("afp")    or "").strip() or None
                    isr    = str(emp.get("isr")    or "").strip() or None
                    personal = str(emp.get("prestamo_personal") or emp.get("personal_loan") or "").strip() or None
                    bank     = str(emp.get("prestamo_bancario") or emp.get("bank_loan")     or "").strip() or None
                    fsv    = str(emp.get("fsv")    or "").strip() or None

                    # ── Departamento / días vacación ─────────────────────────
                    departamento = str(emp.get("departamento") or "").strip() or None
                    dias_vac = emp.get("dias_vacacion_pendientes")
                    try:
                        dias_vac = float(dias_vac) if dias_vac is not None else 0.0
                    except (ValueError, TypeError):
                        dias_vac = 0.0

                    # ── Upsert ───────────────────────────────────────────────
                    cur.execute("SELECT codigo FROM empleados WHERE codigo = ?", (codigo,))
                    exists_by_codigo = cur.fetchone()

                    # If DUI given and codigo not found, check by DUI to avoid duplicates
                    if not exists_by_codigo and dui:
                        cur.execute("SELECT codigo FROM empleados WHERE dui = ?", (dui,))
                        row_by_dui = cur.fetchone()
                        if row_by_dui:
                            # Update by DUI
                            cur.execute(
                                """
                                UPDATE empleados
                                SET nombre=?, cargo=?, fecha_ingreso=?, salario_mensual=?,
                                    dui=?, fecha_fin=?, salario_letras=?, isss=?, afp=?, isr=?,
                                    prestamo_personal=?, prestamo_bancario=?, fsv=?,
                                    departamento=?, fecha_contratacion=?, dias_vacacion_pendientes=?
                                WHERE dui=?
                                """,
                                (nombre, cargo, ingreso, salario, dui, fin, salario_letras, isss, afp, isr,
                                 personal, bank, fsv, departamento, contratacion, dias_vac, dui)
                            )
                            imported_count += 1
                            continue

                    if exists_by_codigo:
                        cur.execute(
                            """
                            UPDATE empleados
                            SET nombre=?, cargo=?, fecha_ingreso=?, salario_mensual=?,
                                dui=?, fecha_fin=?, salario_letras=?, isss=?, afp=?, isr=?,
                                prestamo_personal=?, prestamo_bancario=?, fsv=?,
                                departamento=?, fecha_contratacion=?, dias_vacacion_pendientes=?
                            WHERE codigo=?
                            """,
                            (nombre, cargo, ingreso, salario, dui, fin, salario_letras, isss, afp, isr,
                             personal, bank, fsv, departamento, contratacion, dias_vac, codigo)
                        )
                    else:
                        cur.execute(
                            """
                            INSERT OR IGNORE INTO empleados (
                                codigo, nombre, cargo, fecha_ingreso, salario_mensual, activo, dui,
                                fecha_fin, salario_letras, isss, afp, isr, prestamo_personal, prestamo_bancario, fsv,
                                departamento, fecha_contratacion, dias_vacacion_pendientes
                            ) VALUES (?, ?, ?, ?, ?, 1, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            """,
                            (codigo, nombre, cargo, ingreso, salario, dui, fin, salario_letras, isss, afp, isr,
                             personal, bank, fsv, departamento, contratacion, dias_vac)
                        )
                    imported_count += 1
                except Exception as ex:
                    print(f"[init_db] Error processing JSON employee row: {ex}")
            conn.commit()
            print(f"[init_db] {imported_count} empleados cargados/actualizados desde {mvp_json_path}")
        except Exception as e:
            print(f"[init_db] Error al leer seed_mvp_employees.json: {e}")

    # Seed/update employees from Control de vacaciones.xlsx
    excel_path = r"C:\Users\HP\Documents\Control de vacaciones.xlsx"
    if os.path.exists(excel_path):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            if "Empleados" in wb.sheetnames:
                sheet = wb["Empleados"]
                imported_count = 0
                for idx in range(5, sheet.max_row + 1):
                    row = [cell.value for cell in sheet[idx]]
                    if not row or len(row) < 9 or row[1] is None:
                        continue
                    
                    try:
                        raw_code = row[1]
                        codigo = str(int(raw_code)).zfill(4) if isinstance(raw_code, (int, float)) else str(raw_code).strip().zfill(4)
                        
                        nombre = str(row[2]).strip()
                        
                        ingreso_dt = row[3]
                        if isinstance(ingreso_dt, datetime):
                            fecha_ingreso = ingreso_dt.strftime("%Y-%m-%d")
                        elif ingreso_dt:
                            fecha_ingreso = str(ingreso_dt).strip()[:10]
                        else:
                            fecha_ingreso = ""
                            
                        contratacion_dt = row[4]
                        if isinstance(contratacion_dt, datetime):
                            fecha_contratacion = contratacion_dt.strftime("%Y-%m-%d")
                        elif contratacion_dt:
                            fecha_contratacion = str(contratacion_dt).strip()[:10]
                        else:
                            fecha_contratacion = ""
                            
                        salario = float(row[5]) if row[5] is not None else 0.0
                        departamento = str(row[6]).strip() if row[6] else ""
                        cargo = str(row[7]).strip() if row[7] else ""
                        
                        estado = str(row[8]).strip().lower() if row[8] else "activo"
                        activo = 1 if estado in ("activo", "active", "1") else 0
                        
                        dias_vac = float(row[14]) if (len(row) > 14 and row[14] is not None) else 0.0
                        
                        cur.execute("SELECT codigo FROM empleados WHERE codigo = ?", (codigo,))
                        if cur.fetchone():
                            cur.execute(
                                """
                                UPDATE empleados
                                SET nombre=?, cargo=?, fecha_ingreso=?, salario_mensual=?,
                                    activo=?, departamento=?, fecha_contratacion=?, dias_vacacion_pendientes=?
                                WHERE codigo=?
                                """,
                                (nombre, cargo, fecha_ingreso, salario, activo, departamento, fecha_contratacion, dias_vac, codigo)
                            )
                        else:
                            cur.execute(
                                """
                                INSERT INTO empleados (
                                    codigo, nombre, cargo, fecha_ingreso, salario_mensual, activo,
                                    departamento, fecha_contratacion, dias_vacacion_pendientes
                                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                                """,
                                (codigo, nombre, cargo, fecha_ingreso, salario, activo, departamento, fecha_contratacion, dias_vac)
                            )
                        imported_count += 1
                    except Exception as ex:
                        print(f"[init_db] Error procesando fila {idx} del Excel: {ex}")
                wb.close()
                conn.commit()
                print(f"[init_db] {imported_count} empleados cargados/actualizados desde {excel_path}")
        except Exception as e:
            print(f"[init_db] Error al leer Control de vacaciones.xlsx: {e}")

    # Seed employees from SEED_CSV only if table is empty (fallback)
    cur.execute("SELECT COUNT(*) AS c FROM empleados")
    if cur.fetchone()["c"] == 0 and os.path.exists(SEED_CSV):
        with open(SEED_CSV, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            rows = [
                (r["codigo"], r["nombre"], r["cargo"], r["fecha_ingreso"], float(r["salario_mensual"]))
                for r in reader
            ]
        cur.executemany(
            "INSERT OR IGNORE INTO empleados (codigo, nombre, cargo, fecha_ingreso, salario_mensual) VALUES (?,?,?,?,?)",
            rows,
        )
        conn.commit()
        print(f"[init_db] {len(rows)} empleados cargados desde {SEED_CSV}")

    # Seed a default admin user only if no admin exists
    cur.execute("SELECT COUNT(*) AS c FROM admin_users")
    if cur.fetchone()["c"] == 0:
        default_user = os.environ.get("ADMIN_USER", "admin")
        default_pass = os.environ.get("ADMIN_PASS", "dongbu2026")
        try:
            cur.execute(
                "INSERT OR IGNORE INTO admin_users (username, password_hash, nombre_completo) VALUES (?,?,?)",
                (default_user, generate_password_hash(default_pass), "Administrador RRHH"),
            )
            conn.commit()
            print(f"[init_db] Usuario administrador creado -> usuario: {default_user} / clave: {default_pass}")
            print("[init_db] IMPORTANTE: cambia esta clave despues del primer ingreso.")
        except sqlite3.IntegrityError:
            pass

    # Seed initial registros (incapacidades y vacaciones) if table is empty
    cur.execute("SELECT COUNT(*) AS c FROM registros")
    if cur.fetchone()["c"] == 0:
        registros_json_path = os.path.join(os.path.dirname(__file__), "seed_registros.json")
        if os.path.exists(registros_json_path):
            try:
                import json
                with open(registros_json_path, "r", encoding="utf-8") as f:
                    regs = json.load(f)
                
                imported_regs = 0
                for r in regs:
                    dui = r.get("dui")
                    tipo = r.get("tipo")
                    fi = r.get("fecha_inicio")
                    ff = r.get("fecha_fin")
                    dias = r.get("dias")
                    foto = r.get("foto_path")
                    obs = r.get("observaciones")
                    fecha_reg = r.get("fecha_registro")
                    estado = r.get("estado", "aprobado").lower()
                    
                    # Resolve employee code from DUI
                    cur.execute("SELECT codigo FROM empleados WHERE dui = ?", (dui,))
                    emp_row = cur.fetchone()
                    if emp_row:
                        codigo = emp_row["codigo"]
                    else:
                        # Try matching by name
                        name = r.get("nombre")
                        if name:
                            cur.execute("SELECT codigo FROM empleados WHERE UPPER(nombre) = ?", (name.upper().strip(),))
                            emp_row = cur.fetchone()
                            if emp_row:
                                codigo = emp_row["codigo"]
                            else:
                                continue
                        else:
                            continue
                            
                    cur.execute(
                        """INSERT INTO registros (codigo_empleado, tipo, fecha_inicio, fecha_fin, dias, foto_path, observaciones, fecha_registro, estado)
                           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                        (codigo, tipo, fi, ff, dias, foto, obs, fecha_reg, estado)
                    )
                    imported_regs += 1
                conn.commit()
                print(f"[init_db] {imported_regs} registros de incapacidades y vacaciones cargados.")
            except Exception as e:
                print(f"[init_db] Error seeding registros: {e}")

    # Seed initial templates
    TEMPLATES_DIR = os.path.join(os.path.dirname(DB_PATH), "..", "app", "uploads", "templates")
    seed_root = os.path.join(os.path.dirname(__file__), "seed_templates")
    seeds = [
        ("salario", "Constancia de salario Firmada.docx"),
        ("laboral", "Constancia Laboral.docx"),
    ]
    for doc_type, filename in seeds:
        cur.execute("SELECT id FROM templates WHERE doc_type = ? AND is_active = 1", (doc_type,))
        if not cur.fetchone():
            candidate_1 = os.path.join(seed_root, doc_type, filename)
            candidate_2 = os.path.join(r"C:\Users\HP\Documents\Formato de constancia", filename)
            
            src_path = None
            if os.path.exists(candidate_1):
                src_path = candidate_1
            elif os.path.exists(candidate_2):
                src_path = candidate_2
                
            if src_path:
                folder = os.path.join(TEMPLATES_DIR, doc_type)
                os.makedirs(folder, exist_ok=True)
                target = os.path.join(folder, filename)
                try:
                    shutil.copy2(src_path, target)
                    cur.execute(
                        "INSERT INTO templates (doc_type, filename, stored_path, is_active, created_at) VALUES (?, ?, ?, 1, ?)",
                        (doc_type, filename, target, datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
                    )
                    conn.commit()
                    print(f"[init_db] Plantilla {doc_type} precargada con éxito.")
                except Exception as e:
                    print(f"[init_db] Error al precargar plantilla {doc_type}: {e}")

    conn.close()
